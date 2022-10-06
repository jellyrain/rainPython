import re
from openpyxl import Workbook, load_workbook
from content.lexical_analysis import Lexical_Analysis


class Excel:
    def __init__(self, excel_path: str = None, data_only: bool = True, keep_links: bool = True,
                 read_only=False) -> None:
        self.sheet = None
        if excel_path is None:
            self.excel = Workbook()
        else:
            self.excel = load_workbook(excel_path, read_only, data_only, keep_links)
        self.sheet_index(0)

    def sheetnames(self) -> list[str]:
        """ 获取全部 excel sheet（表） """
        return self.excel.sheetnames

    def sheet_index(self, index: int = 1):
        """ 设置当前操作的 sheet（表） """
        self.sheet = self.excel[self.excel.sheetnames[index]]
        return self

    def sheet_str(self, sheet_name: str):
        """ 设置当前操作的 sheet（表） """
        self.sheet = self.excel[sheet_name]
        return self

    def copy(self, sheet_name: str, new_sheet_name: str = None):
        """ 复制一个工作表 并且 将当前操作的表 设置为此表"""
        self.sheet = self.excel.copy_worksheet(self.excel[sheet_name])
        if new_sheet_name:
            self.sheet.title = new_sheet_name
        return self

    def remove(self, sheet_str: str, sheet_index: int = 0):
        """ 删除表 并且将当前操作的表 默认重置为第一张表 """
        self.excel.remove(sheet_str)
        self.sheet_index(sheet_index)
        return self

    def save(self, file_name: str):
        """ 保存 excel 工作簿 """
        self.excel.save(file_name)
        return self

    def delete_row(self, index: int, amount: int = 1):
        """ 删除一行或者多行 """
        self.sheet.delete_rows(index, amount)
        return self

    def delete_col(self, index: int, amount: int = 1):
        """ 删除一列或者多列 """
        self.sheet.delete_cols(index, amount)
        return self

    def row(self, start_row: int = 1, end_row: int = -1) -> list:
        """ 获取指定行 使用二维数组 内维一个数组就是一行 """
        arr = []
        # 判断是否使用最大行数
        if end_row == -1:
            end_row = self.sheet.max_row
        # 读取值并且存入 r 数组中 并且 去掉 全 None 数组 放到 arr 中
        for i in range(start_row, end_row + 1):
            r = []
            for j in range(1, self.sheet.max_column + 1):
                r.append(self.sheet.cell(i, j).value)
            if any(r):
                arr.append(r)

        return arr

    def column(self, start_column: int = 1, end_column: int = -1) -> list:
        """ 获取指定列 使用二维数组 内维一个数组就是一列 """
        arr = []
        # 判断是否使用最大列数
        if end_column == -1:
            end_column = self.sheet.max_column
        # 读取值并且存入 r 数组中 并且 去掉 全 None 数组 放到 arr 中
        for i in range(start_column, end_column + 1):
            r = []
            for j in range(1, self.sheet.max_row + 1):
                r.append(self.sheet.cell(j, i).value)
            arr.append(r)

        return arr


class Intercept_Excel:
    def __init__(self, excel_path: str = None, data_only: bool = True, keep_links: bool = True,
                 read_only=False) -> None:
        self.__load_excel = Excel(excel_path, data_only, keep_links, read_only)
        self.__excel = Workbook()
        self.__sheet = self.__excel[self.__excel.sheetnames[0]]
        self.__column_number = 0

    def save(self, file_name: str):
        """ 保存 excel 工作簿 """
        self.__excel.save(file_name)
        return self

    def get_column_number(self) -> int:
        """ 获取当前表中有多少列 """
        return self.__column_number

    def add_column_base(self, column: int, func, is_not_none: bool = False):
        """
        添加 列 基 方法 

        column: 第几列

        func: 筛选值的函数: func(item: str) -> str

        is_not_none: 是否排除 None 行
        """
        column_value = self.__load_excel.column(column, column)[0]
        self.__column_number += 1
        for index, item in enumerate(column_value):
            if is_not_none and item is None:
                continue
            self.__sheet.cell(index + 1, self.__column_number, func(item))
        return self

    def add_column(self, column: int, is_not_none: bool = False):
        """ 不筛选取列 """
        self.add_column_base(column, lambda item: item, is_not_none)
        return self

    def add_column_regexp(self, column: int, pattern: str, is_not_none: bool = False):
        """ 取列正则筛选值 """

        def func(item: str) -> str:
            if item is None:
                return item
            result = re.findall(pattern, item)
            if len(result) == 0:
                return None
            return result[0]

        self.add_column_base(column, func, is_not_none)
        return self

    def add_colum_if_semanteme(self, column: int, str_list: list[str], weight: float = 0.7, is_not_none: bool = False):
        """ 和字符串列表对比语义类似筛选替换值 """

        def func(item: str) -> str | None:
            if item is None: return None
            return Lexical_Analysis(item).list_extractOne(str_list, weight)

        self.add_column_base(column, func, is_not_none)
        return self


__all__ = ['Excel', 'Intercept_Excel']
